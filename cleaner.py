"""
Excel 数据清洗能力模块
======================
提供"读取 Excel → 清洗数据 → 生成结果 Excel"的全部能力。
本模块不依赖 Streamlit，可独立运行和测试。

使用方式:
    import cleaner
    result_wb, metrics = cleaner.clean("报名数据.xlsx")
    result_wb.save("数据清洗结果.xlsx")
    print(metrics)
"""

import io
import re
from datetime import datetime, date
from typing import Dict, List, Optional, Tuple
from copy import copy

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

# ============================================================
# SECTION 1: 常量与配置
# ============================================================

# —— 工作表名称 ——
SHEET_NAME = "报名明细"

# —— 必要列（上传的 Excel 必须包含这些列） ——
REQUIRED_COLUMNS = [
    "姓名",
    "手机号",
    "部门",
    "城市",
    "报名课程",
    "报名日期",
    "报名状态",
    "应缴金额",
    "实缴金额",
    "备注",
]

# —— 中国大陆手机号正则 ——
PHONE_PATTERN = re.compile(r"^1[3-9]\d{9}$")

# —— 默认报名状态 ——
DEFAULT_STATUS = "待确认"

# —— 部门名称标准化映射 ——
# 键：可能出现的各种写法（小写去除空格），值：标准名称
DEPARTMENT_MAP = {
    "技术部": "技术部",
    "技术部门": "技术部",
    "技术中心": "技术部",
    "技术研发部": "技术部",
    "tech": "技术部",
    "研发部": "研发部",
    "研发中心": "研发部",
    "研发部门": "研发部",
    "产品研发部": "研发部",
    "市场部": "市场部",
    "市场部门": "市场部",
    "市场营销": "市场部",
    "市场营销部": "市场部",
    "marketing": "市场部",
    "销售部": "销售部",
    "销售部门": "销售部",
    "业务部": "销售部",
    "sales": "销售部",
    "人力资源部": "人力资源部",
    "人事部": "人力资源部",
    "人资部": "人力资源部",
    "hr": "人力资源部",
    "财务部": "财务部",
    "财务部门": "财务部",
    "finance": "财务部",
    "行政部": "行政部",
    "行政部门": "行政部",
    "综合部": "行政部",
    "综合管理部": "行政部",
    "admin": "行政部",
}

# —— 课程名称标准化映射 ——
COURSE_MAP = {
    "python入门": "Python入门",
    "python基础": "Python入门",
    "python 入门": "Python入门",
    "python初级": "Python入门",
    "excel进阶": "Excel进阶",
    "excel高级": "Excel进阶",
    "excel 进阶": "Excel进阶",
    "excel高级应用": "Excel进阶",
    "数据分析": "数据分析",
    "数据分析实战": "数据分析",
    "数据分析基础": "数据分析",
    "数据可视化": "数据分析",
    "项目管理": "项目管理",
    "项目管理实战": "项目管理",
    "项目实战": "项目管理",
    "pm": "项目管理",
    "ai实战": "AI实战",
    "ai 实战": "AI实战",
    "人工智能实战": "AI实战",
    "ai应用": "AI实战",
    "云计算基础": "云计算基础",
    "云计算": "云计算基础",
    "cloud基础": "云计算基础",
    "网络攻防": "网络攻防",
    "网络安全": "网络攻防",
    "web安全": "网络攻防",
    "产品经理": "产品经理",
    "产品经理实战": "产品经理",
    "产品管理": "产品经理",
    "ui设计": "UI设计",
    "ui设计基础": "UI设计",
    "用户体验设计": "UI设计",
    "ux设计": "UI设计",
}

# —— 样式定义 ——
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="微软雅黑")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

DATA_ALIGNMENT = Alignment(vertical="center")
NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")


# ============================================================
# SECTION 2: 基础工具函数
# ============================================================

def _read_headers(ws) -> List[str]:
    """读取工作表第一行作为表头，None 值替换为 '列N'。"""
    headers = []
    for idx, cell in enumerate(ws[1], start=1):
        headers.append(str(cell.value).strip() if cell.value is not None else f"列{idx}")
    return headers


def _find_column_index(headers: List[str], col_name: str) -> int:
    """
    在表头中精确查找列名，返回 1-based 列号。
    找不到时返回 -1。
    """
    for idx, h in enumerate(headers, start=1):
        if h == col_name:
            return idx
    return -1


def _is_empty(value) -> bool:
    """判断单元格值是否为空（None、空字符串、纯空格）。"""
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _normalize_key(value) -> str:
    """
    将值转为标准化键（小写、去空格），用于模糊匹配。
    例如 '  Python 入门  ' → 'python入门'
    """
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "")


def _count_data_rows(ws) -> int:
    """统计数据行数（不含表头）。"""
    if ws.max_row <= 1:
        return 0
    return ws.max_row - 1


def _style_header(ws):
    """为工作表表头设置统一样式。"""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _style_data_cell(cell, col_type: str = "text"):
    """为数据单元格设置样式。"""
    cell.border = THIN_BORDER
    if col_type == "number":
        cell.alignment = NUMBER_ALIGNMENT
        if isinstance(cell.value, (int, float)):
            cell.number_format = "#,##0.00"
    else:
        cell.alignment = DATA_ALIGNMENT


def _auto_width(ws):
    """自动调整列宽（根据内容估算）。"""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                # 中文字符算 2 个宽度
                val = str(cell.value)
                length = 0
                for ch in val:
                    length += 2 if ord(ch) > 127 else 1
                max_length = max(max_length, length)
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)


# ============================================================
# SECTION 3: 6 条清洗规则
# ============================================================

def _rule_trim_names(ws, name_col: int) -> int:
    """
    规则1：清理姓名前后空格。
    对姓名列的每个单元格执行 strip()，统计被修改的数量。
    """
    changed = 0
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=name_col)
        if isinstance(cell.value, str):
            stripped = cell.value.strip()
            if stripped != cell.value:
                cell.value = stripped
                changed += 1
    return changed


def _rule_standardize_department(ws, dept_col: int) -> int:
    """
    规则2：统一部门名称。
    根据 DEPARTMENT_MAP 将变体写法映射到标准名称。
    无法匹配的保留原样。
    """
    changed = 0
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=dept_col)
        if _is_empty(cell.value):
            continue
        key = _normalize_key(cell.value)
        if key in DEPARTMENT_MAP:
            standard = DEPARTMENT_MAP[key]
            if standard != cell.value:
                cell.value = standard
                changed += 1
    return changed


def _rule_standardize_course(ws, course_col: int) -> int:
    """
    规则3：统一课程名称。
    根据 COURSE_MAP 将变体写法映射到标准名称。
    """
    changed = 0
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=course_col)
        if _is_empty(cell.value):
            continue
        key = _normalize_key(cell.value)
        if key in COURSE_MAP:
            standard = COURSE_MAP[key]
            if standard != cell.value:
                cell.value = standard
                changed += 1
    return changed


def _rule_fill_status(ws, status_col: int) -> int:
    """
    规则4：补全空报名状态。
    对报名状态为空的行，填入默认值"待确认"。
    """
    filled = 0
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=status_col)
        if _is_empty(cell.value):
            cell.value = DEFAULT_STATUS
            filled += 1
    return filled


def _rule_check_phone(ws, phone_col: int) -> List[Dict]:
    """
    规则5：检查手机号异常。
    - 转为字符串，去除非数字字符
    - 验证是否为合法手机号
    - 不合法的用红色标记，记录异常

    返回异常记录列表，每条包含 {行号, 原始值, 问题描述}。
    """
    anomalies = []
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=phone_col)
        original = cell.value

        if _is_empty(original):
            cell.fill = ERROR_FILL
            anomalies.append({
                "行号": row_idx - 1,  # 数据行号（不含表头）
                "字段": "手机号",
                "原始值": str(original) if original else "(空)",
                "问题描述": "手机号为空",
            })
            continue

        # 去除非数字字符
        digits = re.sub(r"\D", "", str(original))

        # 去掉中国国家代码前缀
        if len(digits) == 12 and digits.startswith("86"):
            digits = digits[2:]
        if len(digits) == 13 and digits.startswith("86"):
            digits = digits[2:]

        # 验证合法性
        if not PHONE_PATTERN.match(digits):
            cell.fill = ERROR_FILL
            problem = (
                f"手机号格式异常（'{original}'→'{digits}'，"
                f"{'位数不对' if len(digits) != 11 else '号段不合法'}）"
            )
            anomalies.append({
                "行号": row_idx - 1,
                "字段": "手机号",
                "原始值": str(original),
                "问题描述": problem,
            })
        else:
            # 合法：保存纯数字格式
            if digits != str(original):
                cell.value = digits
                cell.number_format = "@"

    return anomalies


def _rule_find_duplicates(ws, phone_col: int, name_col: int) -> List[Dict]:
    """
    规则6：识别重复报名。
    按手机号分组，同一手机号出现 ≥2 次视为重复报名。
    用黄色标记重复行（保留第一次出现的行不标记）。

    返回重复记录列表。
    """
    # 第一遍：统计每个手机号出现的行号
    phone_rows: Dict[str, List[int]] = {}
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=phone_col)
        digits = re.sub(r"\D", "", str(cell.value)) if cell.value else ""
        if PHONE_PATTERN.match(digits):
            phone_rows.setdefault(digits, []).append(row_idx)

    # 第二遍：标记重复行（跳过第一个，标记其余的）
    duplicates = []
    for digits, rows in phone_rows.items():
        if len(rows) >= 2:
            for row_idx in rows[1:]:  # 保留第一个，标记后续
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col).fill = WARN_FILL
                name = ws.cell(row=row_idx, column=name_col).value
                first_row = rows[0]
                first_name = ws.cell(row=first_row, column=name_col).value
                duplicates.append({
                    "行号": row_idx - 1,
                    "字段": "手机号",
                    "原始值": digits,
                    "问题描述": (
                        f"手机号 {digits} 重复报名（"
                        f"当前：第{row_idx - 1}行「{name}」，"
                        f"首次：第{first_row - 1}行「{first_name}」）"
                    ),
                })

    return duplicates


# ============================================================
# SECTION 4: 主清洗流程
# ============================================================

def clean(file_path_or_bytes, enabled_rules: dict = None) -> Tuple[openpyxl.Workbook, Dict]:
    """
    执行完整的数据清洗流程。

    参数:
        file_path_or_bytes: Excel 文件路径（str），或文件字节流（bytes）
        enabled_rules: 可选，{'rule_trim_names': True, 'rule_dept': False, ...}
                      为 None 时默认启用全部规则

    返回:
        (result_wb, metrics)
        - result_wb: openpyxl.Workbook，包含「清洗结果」「汇总报表」「异常记录」三个工作表
        - metrics: dict，包含 5 个摘要指标

    异常:
        ValueError: 工作表不存在、必要列缺失
        FileNotFoundError: 文件路径不存在
        openpyxl.utils.exceptions.InvalidFileException: 文件格式不支持
    """
    # ============================================================
    # 4a: 加载文件
    # ============================================================
    if isinstance(file_path_or_bytes, bytes):
        source_wb = openpyxl.load_workbook(
            io.BytesIO(file_path_or_bytes), data_only=True
        )
        source_is_file = False
    elif isinstance(file_path_or_bytes, str):
        source_wb = openpyxl.load_workbook(
            file_path_or_bytes, data_only=True
        )
        source_is_file = True
    else:
        raise TypeError("file_path_or_bytes 必须是文件路径(str)或字节流(bytes)")

    # ============================================================
    # 4b: 检查工作表是否存在
    # ============================================================
    if SHEET_NAME not in source_wb.sheetnames:
        available = "、".join(source_wb.sheetnames)
        raise ValueError(
            f"❌ 未找到工作表「{SHEET_NAME}」。\n"
            f"   当前文件包含的工作表: {available}\n"
            f"   请确认工作表名称是否正确。"
        )

    ws = source_wb[SHEET_NAME]
    headers = _read_headers(ws)

    # ============================================================
    # 4c: 检查必要列是否存在
    # ============================================================
    missing_columns = []
    column_indices = {}  # 列名 → 列号（1-based）
    for col_name in REQUIRED_COLUMNS:
        idx = _find_column_index(headers, col_name)
        if idx == -1:
            missing_columns.append(col_name)
        else:
            column_indices[col_name] = idx

    if missing_columns:
        raise ValueError(
            f"❌ 缺少必要列: {'、'.join(missing_columns)}\n"
            f"   当前表头: {'、'.join(headers)}\n"
            f"   必要列清单: {'、'.join(REQUIRED_COLUMNS)}"
        )

    total_records = _count_data_rows(ws)
    if total_records == 0:
        raise ValueError("❌ 工作表中没有数据行（仅包含表头），无法清洗。")

    # ============================================================
    # 4d: 直接使用已加载的工作簿（数据源已在内存中，修改不影响原始文件）
    # ============================================================
    cleaned_wb = source_wb
    cleaned_ws = ws
    cleaned_cols = column_indices

    # ============================================================
    # 4e: 执行 6 条清洗规则，收集异常
    # ============================================================
    all_anomalies: List[Dict] = []

    # 辅助：判断规则是否启用（未传 enabled_rules 时默认全部启用）
    def _enabled(rule_id: str) -> bool:
        return enabled_rules is None or enabled_rules.get(rule_id, True)

    # 规则1: 清理姓名前后空格
    trim_count = _rule_trim_names(cleaned_ws, cleaned_cols["姓名"]) if _enabled("rule_trim_names") else 0

    # 规则2: 统一部门名称
    dept_count = _rule_standardize_department(cleaned_ws, cleaned_cols["部门"]) if _enabled("rule_dept") else 0

    # 规则3: 统一课程名称
    course_count = _rule_standardize_course(cleaned_ws, cleaned_cols["报名课程"]) if _enabled("rule_course") else 0

    # 规则4: 补全空报名状态
    status_count = _rule_fill_status(cleaned_ws, cleaned_cols["报名状态"]) if _enabled("rule_status") else 0

    # 规则5: 检查手机号异常
    if _enabled("rule_phone"):
        phone_anomalies = _rule_check_phone(cleaned_ws, cleaned_cols["手机号"])
        all_anomalies.extend(phone_anomalies)
    else:
        phone_anomalies = []

    # 规则6: 识别重复报名（基于手机号）
    if _enabled("rule_duplicate"):
        dup_records = _rule_find_duplicates(
            cleaned_ws, cleaned_cols["手机号"], cleaned_cols["姓名"]
        )
        all_anomalies.extend(dup_records)
    else:
        dup_records = []

    # ============================================================
    # 4f: 计算汇总指标
    # ============================================================
    amount_due_total = 0.0  # 应缴金额合计
    amount_paid_total = 0.0  # 实缴金额合计

    for row_idx in range(2, cleaned_ws.max_row + 1):
        due_cell = cleaned_ws.cell(row=row_idx, column=cleaned_cols["应缴金额"])
        paid_cell = cleaned_ws.cell(row=row_idx, column=cleaned_cols["实缴金额"])

        try:
            amount_due_total += float(due_cell.value) if due_cell.value is not None else 0
        except (ValueError, TypeError):
            pass
        try:
            amount_paid_total += float(paid_cell.value) if paid_cell.value is not None else 0
        except (ValueError, TypeError):
            pass

    # 重复手机号数 = 出现次数 ≥2 的不同手机号个数
    duplicate_phone_count = len(set(
        a["原始值"] for a in dup_records
    ))

    # 异常记录数 = 手机号异常数 + 重复报名记录数
    anomaly_count = len(all_anomalies)

    metrics = {
        "总记录数": total_records,
        "异常记录数": anomaly_count,
        "重复手机号数": duplicate_phone_count,
        "应缴金额合计": round(amount_due_total, 2),
        "实缴金额合计": round(amount_paid_total, 2),
    }

    # ============================================================
    # 4g: 构建结果 Excel（3个工作表）
    # ============================================================
    result_wb = openpyxl.Workbook()

    # —— 删除默认创建的空白工作表 ——
    result_wb.remove(result_wb.active)

    _build_cleaned_sheet(result_wb, cleaned_ws, cleaned_headers, cleaned_cols)
    _build_summary_sheet(result_wb, metrics, trim_count, dept_count,
                          course_count, status_count, phone_anomalies, dup_records)
    _build_anomaly_sheet(result_wb, all_anomalies)

    return result_wb, metrics


# ============================================================
# SECTION 5: 结果 Excel 的三个工作表构建
# ============================================================

def _build_cleaned_sheet(result_wb, cleaned_ws, headers, col_indices):
    """
    工作表1: 清洗结果。
    将清洗后的全部数据复制过来，保留样式（异常行的颜色标记）。
    """
    ws_out = result_wb.create_sheet("清洗结果")

    # 复制表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws_out.cell(row=1, column=col_idx, value=header)
    _style_header(ws_out)

    # 复制数据行（保留已设置的 fill 样式）
    for row_idx in range(2, cleaned_ws.max_row + 1):
        for col_idx in range(1, cleaned_ws.max_column + 1):
            src_cell = cleaned_ws.cell(row=row_idx, column=col_idx)
            dst_cell = ws_out.cell(row=row_idx, column=col_idx, value=src_cell.value)
            dst_cell.border = THIN_BORDER
            dst_cell.alignment = DATA_ALIGNMENT
            # 保留异常标记的填充色
            if src_cell.fill and src_cell.fill.start_color and src_cell.fill.start_color.rgb:
                rgb = src_cell.fill.start_color.rgb
                if rgb in ("00FFC7CE", "FFC7CE", "00FFEB9C", "FFEB9C"):
                    dst_cell.fill = copy(src_cell.fill)

    _auto_width(ws_out)
    ws_out.freeze_panes = "A2"


def _build_summary_sheet(result_wb, metrics, trim_count, dept_count,
                          course_count, status_count, phone_anomalies, dup_records):
    """
    工作表2: 汇总报表。
    包含清洗指标卡片和清洗规则执行详情。
    """
    ws_out = result_wb.create_sheet("汇总报表")

    # 标题
    ws_out.merge_cells("A1:C1")
    title_cell = ws_out.cell(row=1, column=1, value="数据清洗汇总报表")
    title_cell.font = Font(bold=True, size=16, name="微软雅黑")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # 指标卡片区域
    metric_labels = [
        ("总记录数", metrics["总记录数"], "条"),
        ("异常记录数", metrics["异常记录数"], "条"),
        ("重复手机号数", metrics["重复手机号数"], "个"),
        ("应缴金额合计", metrics["应缴金额合计"], "元"),
        ("实缴金额合计", metrics["实缴金额合计"], "元"),
    ]

    row = 3
    ws_out.cell(row=row, column=1, value="📊 核心指标").font = Font(bold=True, size=13, name="微软雅黑")
    row += 1

    # 表头
    for col, label in enumerate(["指标", "数值", "单位"], 1):
        cell = ws_out.cell(row=row, column=col, value=label)
        cell.font = Font(bold=True, color="FFFFFF", name="微软雅黑")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = HEADER_ALIGNMENT
    row += 1

    for label, value, unit in metric_labels:
        ws_out.cell(row=row, column=1, value=label).border = THIN_BORDER
        val_cell = ws_out.cell(row=row, column=2, value=value)
        val_cell.border = THIN_BORDER
        val_cell.alignment = NUMBER_ALIGNMENT
        if isinstance(value, float):
            val_cell.number_format = "#,##0.00"
        ws_out.cell(row=row, column=3, value=unit).border = THIN_BORDER

        # 异常指标用红色高亮
        if label in ("异常记录数", "重复手机号数") and value > 0:
            val_cell.fill = ERROR_FILL
        row += 1

    # 清洗规则执行详情
    row += 1
    ws_out.cell(row=row, column=1, value="🔧 清洗规则执行详情").font = Font(bold=True, size=13, name="微软雅黑")
    row += 1

    rule_headers = ["规则", "处理数量", "说明"]
    for col, label in enumerate(rule_headers, 1):
        cell = ws_out.cell(row=row, column=col, value=label)
        cell.font = Font(bold=True, color="FFFFFF", name="微软雅黑")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = HEADER_ALIGNMENT
    row += 1

    rules_detail = [
        ("清理姓名前后空格", trim_count,
         "修改了" + (f"{trim_count} 个姓名" if trim_count > 0 else "0 个（姓名已整洁）")),
        ("统一部门名称", dept_count,
         "标准化了" + (f"{dept_count} 个部门名称" if dept_count > 0 else "0 个（部门名称已统一）")),
        ("统一课程名称", course_count,
         "标准化了" + (f"{course_count} 个课程名称" if course_count > 0 else "0 个（课程名称已统一）")),
        ("补全空报名状态", status_count,
         "补全了" + (f"{status_count} 个空报名状态" if status_count > 0 else "0 个（无空状态）")),
        ("检查手机号异常", len(phone_anomalies),
         "发现" + (f"{len(phone_anomalies)} 个异常手机号" if phone_anomalies else "0 个（手机号均正常）")),
        ("识别重复报名", len(dup_records),
         "识别出" + (f"{len(dup_records)} 条重复报名记录" if dup_records else "0 条（无重复报名）")),
    ]

    for rule_name, count, desc in rules_detail:
        ws_out.cell(row=row, column=1, value=rule_name).border = THIN_BORDER
        cnt_cell = ws_out.cell(row=row, column=2, value=count)
        cnt_cell.border = THIN_BORDER
        cnt_cell.alignment = NUMBER_ALIGNMENT
        ws_out.cell(row=row, column=3, value=desc).border = THIN_BORDER
        if count > 0:
            cnt_cell.fill = WARN_FILL
        row += 1

    _auto_width(ws_out)
    # 设置 A 列稍宽
    ws_out.column_dimensions["A"].width = 22
    ws_out.column_dimensions["C"].width = 38


def _build_anomaly_sheet(result_wb, anomalies):
    """
    工作表3: 异常记录。
    列出所有检测到的问题记录。
    """
    ws_out = result_wb.create_sheet("异常记录")

    # 标题
    ws_out.merge_cells("A1:D1")
    title_cell = ws_out.cell(row=1, column=1, value="异常记录明细")
    title_cell.font = Font(bold=True, size=16, name="微软雅黑")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # 表头
    anomaly_headers = ["序号", "数据行号", "异常字段", "问题描述"]
    row = 3
    for col, label in enumerate(anomaly_headers, 1):
        cell = ws_out.cell(row=row, column=col, value=label)
        cell.font = Font(bold=True, color="FFFFFF", name="微软雅黑")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = HEADER_ALIGNMENT
    row += 1

    if anomalies:
        for idx, a in enumerate(anomalies, 1):
            ws_out.cell(row=row, column=1, value=idx).border = THIN_BORDER
            ws_out.cell(row=row, column=2, value=a.get("行号", "")).border = THIN_BORDER
            ws_out.cell(row=row, column=3, value=a.get("字段", "")).border = THIN_BORDER
            ws_out.cell(row=row, column=4, value=a.get("问题描述", "")).border = THIN_BORDER

            # 根据类型着色
            desc = a.get("问题描述", "")
            if "重复" in desc:
                for c in range(1, 5):
                    ws_out.cell(row=row, column=c).fill = WARN_FILL
            else:
                for c in range(1, 5):
                    ws_out.cell(row=row, column=c).fill = ERROR_FILL
            row += 1
    else:
        ws_out.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws_out.cell(row=row, column=1, value="🎉 未发现任何异常记录，数据质量良好！").alignment = Alignment(horizontal="center")
        for c in range(1, 5):
            ws_out.cell(row=row, column=c).fill = GOOD_FILL
            ws_out.cell(row=row, column=c).border = THIN_BORDER

    _auto_width(ws_out)
    ws_out.column_dimensions["A"].width = 8
    ws_out.column_dimensions["B"].width = 12
    ws_out.column_dimensions["C"].width = 14
    ws_out.column_dimensions["D"].width = 60


# ============================================================
# SECTION 6: 便捷函数
# ============================================================

def clean_and_save(input_path: str, output_path: str = "数据清洗结果.xlsx",
                   enabled_rules: dict = None) -> Dict:
    """
    一站式函数：读取文件 → 清洗 → 保存结果。

    参数:
        input_path: 输入 Excel 文件路径
        output_path: 输出 Excel 文件路径，默认为「数据清洗结果.xlsx」

    返回:
        metrics: 5 个摘要指标
    """
    result_wb, metrics = clean(input_path, enabled_rules=enabled_rules)
    result_wb.save(output_path)
    return metrics


# ============================================================
# SECTION 7: 命令行入口（方便直接测试）
# ============================================================

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        # 默认使用测试文件
        test_file = "报名数据.xlsx"
        print(f"用法: python cleaner.py <Excel文件路径>")
        print(f"未指定文件，尝试使用默认测试文件: {test_file}")
        if __import__("os").path.exists(test_file):
            input_file = test_file
        else:
            print(f"测试文件 {test_file} 不存在，退出。")
            sys.exit(1)
    else:
        input_file = sys.argv[1]

    try:
        print(f"正在读取: {input_file}")
        result_wb, metrics = clean(input_file)

        output_file = "数据清洗结果.xlsx"
        result_wb.save(output_file)
        print(f"\n[OK] 清洗完成！结果已保存到: {output_file}")
        print(f"\n[*] 清洗摘要:")
        print(f"   总记录数:    {metrics['总记录数']}")
        print(f"   异常记录数:  {metrics['异常记录数']}")
        print(f"   重复手机号数: {metrics['重复手机号数']}")
        print(f"   应缴金额合计: {metrics['应缴金额合计']:,.2f}")
        print(f"   实缴金额合计: {metrics['实缴金额合计']:,.2f}")

    except ValueError as e:
        print(f"\n{e}")
        sys.exit(1)
    except FileNotFoundError:
        print(f"\n[ERROR] 文件不存在: {input_file}")
        sys.exit(1)
    except Exception as e:
        print(f"\n[ERROR] 清洗失败: {e}")
        sys.exit(1)
